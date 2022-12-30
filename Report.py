import re

import openpyxl
from openpyxl.styles import Side, Border, Font
import matplotlib.pyplot as plt
import numpy as np


class Report:
    def __init__(self, salary_levels_by_years: dict,
                 number_vacancies_by_years: dict,
                 salaries_years_chosen_profession: dict,
                 number_vacancies_years_chosen_profession: dict,
                 salary_by_city: dict,
                 share_vacancies_by_city: dict,
                 profession: str):

        self.__salary_levels_by_years = salary_levels_by_years
        self.__number_vacancies_by_years = number_vacancies_by_years
        self.__salaries_years_chosen_profession = salaries_years_chosen_profession
        self.__number_vacancies_years_chosen_profession = number_vacancies_years_chosen_profession
        self.__salary_by_city = salary_by_city
        self.__share_vacancies_by_city = share_vacancies_by_city
        self.__profession = profession

    def __generate_statistic_years(self, book, sheet_name, index):
        book.create_sheet(sheet_name)
        title = ["Год",
                 "Средняя зарплата",
                 f"Средняя зарплата - {self.__profession}",
                 "Количество вакансий",
                 f"Количество вакансий - {self.__profession}"]
        book.worksheets[index].append(title)
        for year in self.__salary_levels_by_years.keys():
            book.worksheets[index].append([year,
                                           self.__salary_levels_by_years[year],
                                           self.__salaries_years_chosen_profession[year],
                                           self.__number_vacancies_by_years[year],
                                           self.__number_vacancies_years_chosen_profession[year]])

        for i in range(len(title)):
            book.worksheets[index].cell(1, i + 1).font = Font(bold=True)

        side = Side(border_style='thin', color="FF000000")
        border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )
        for i in range(len(self.__salary_levels_by_years.keys()) + 1):
            for j in range(len(title)):
                book.worksheets[index].cell(i + 1, j + 1).border = border

        dimensions = {}
        for row in book.worksheets[index].rows:
            for cell in row:
                if cell.value:
                    dimensions[cell.column_letter] = max((dimensions.get(cell.column_letter, 0),
                                                          len(str(cell.value)) + 2))
        for col, value in dimensions.items():
            book.worksheets[index].column_dimensions[col].width = value

    def __generate_statistic_cities(self, book, sheet_name, index):
        book.create_sheet(sheet_name)
        side = Side(border_style='thin', color="FF000000")
        border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )

        book.create_sheet("Статистика по городам")
        title = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]

        book.worksheets[index].append(title)

        cities1 = list(self.__salary_by_city.keys())
        cities2 = list(self.__share_vacancies_by_city.keys())
        for i in range(len(cities1)):
            book.worksheets[index].append([cities1[i],
                                           self.__salary_by_city[cities1[i]],
                                           "",
                                           cities2[i],
                                           self.__share_vacancies_by_city[cities2[i]]])

        for i in range(len(title)):
            book.worksheets[index].cell(1, i + 1).font = Font(bold=True)

        for i in range(len(cities1) + 1):
            for j in range(len(title)):
                book.worksheets[index].cell(i + 1, j + 1).border = border

        for i in range(2, len(cities2) + 2):
            book.worksheets[index].cell(i, 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

        dimensions = {}
        for row in book.worksheets[1].rows:
            for cell in row:
                if cell.value:
                    dimensions[cell.column_letter] = max(
                        (dimensions.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dimensions.items():
            book.worksheets[index].column_dimensions[col].width = value

    def generate_excel(self):
        book = openpyxl.Workbook()
        book.remove(book["Sheet"])
        self.__generate_statistic_years(book, "Статистика по годам", 0)
        self.__generate_statistic_cities(book, "Статистика по городам", 1)
        book.save("report.xlsx")

    def generate_image(self):
        figure, ax = plt.subplots(2, 2)

        width = 0.35

        # 1 график
        labels = list(self.__salary_levels_by_years.keys())
        x = np.arange(len(labels))
        ax[0, 0].bar(x - width / 2, self.__salary_levels_by_years.values(), width, label="средняя з/п")
        ax[0, 0].bar(x + width / 2, self.__salaries_years_chosen_profession.values(), width,
                     label=f"з/п {self.__profession}")

        ax[0, 0].set_title("Уровень зарплат по годам")
        ax[0, 0].set_xticks(x, labels)
        ax[0, 0].legend(fontsize=8)
        ax[0, 0].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 0].get_xticklabels() + ax[0, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 0].grid(axis="y")

        # 2 график
        labels = list(self.__number_vacancies_by_years.keys())
        x = np.arange(len(labels))
        ax[0, 1].bar(x - width / 2, self.__number_vacancies_by_years.values(), width, label="Количество вакансий")
        ax[0, 1].bar(x + width / 2, self.__number_vacancies_years_chosen_profession.values(), width,
                     label=f"Количество вакансий\n{self.__profession}")

        ax[0, 1].set_title("Количество вакансий по годам")
        ax[0, 1].set_xticks(x, labels)
        ax[0, 1].legend(loc="upper left", fontsize=8)
        ax[0, 1].set_xticklabels(labels, rotation=90)
        for label in (ax[0, 1].get_xticklabels() + ax[0, 1].get_yticklabels()):
            label.set_fontsize(8)
        ax[0, 1].grid(axis="y")

        # 3 график
        labels = []
        for city in list(reversed(self.__salary_by_city.keys())):
            labels.append("\n".join(re.split(r"[ -]", city)))
        x = np.arange(len(labels))
        ax[1, 0].barh(x - width / 2, list(reversed(self.__salary_by_city.values())), width)

        ax[1, 0].set_title("Уровень зарплат по городам")
        for label in (ax[1, 0].get_xticklabels() + ax[1, 0].get_yticklabels()):
            label.set_fontsize(8)
        ax[1, 0].set_yticks(x, labels, fontsize=6, horizontalalignment="right", verticalalignment="center")
        ax[1, 0].grid(axis="x")

        # 4 график
        self.__share_vacancies_by_city["Другое"] = 1 - sum(self.__share_vacancies_by_city.values())
        labels = list(self.__share_vacancies_by_city.keys())
        ax[1, 1].pie(self.__share_vacancies_by_city.values(), labels=labels, startangle=90, textprops={"fontsize": 6})
        ax[1, 1].set_title("Доля выкансий по городам")
        ax[1, 1].axis("equal")
        for label in (ax[1, 1].get_label()):
            label.set_fontsize(6)

        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()